"""结构化数据表 API — 导入、列表、预览、穿透搜索。"""

import logging
from datetime import datetime
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from sqlalchemy import and_, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db, get_visible_owner_ids
from app.models.asset import ETLDataSource, ETLSyncState
from app.models.content_entity_link import ContentEntityLink
from app.models.structured_table import StructuredTable, StructuredTableRow
from app.models.tag import ContentTag, TagDefinition
from app.models.user import User
from app.schemas.structured_table import (
    BatchDeleteRequest,
    ImportBitableRequest,
    ImportFromURLRequest,
    ImportSpreadsheetRequest,
    SearchResponse,
    SearchResultItem,
    StructuredTableDetail,
    StructuredTableListResponse,
    StructuredTableOut,
    StructuredTableRowListResponse,
    StructuredTableRowOut,
    URLParseResult,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/structured-tables", tags=["结构化数据表"])


async def _apply_extraction_after_import(
    db: AsyncSession, table_id: int, extraction_rule_id: int
) -> None:
    """导入完成后自动应用提取规则（内部辅助函数）。"""
    from app.services.etl.enricher import extract_key_info

    result = await db.execute(
        select(StructuredTable).where(StructuredTable.id == table_id)
    )
    table_obj = result.scalar_one_or_none()
    if not table_obj:
        logger.warning("表格 %d 不存在，跳过提取规则应用", table_id)
        return

    # 使用 content_text 作为提取内容（整个表格汇总文本）
    content = table_obj.content_text or ""
    if not content and table_obj.summary:
        content = table_obj.summary
    if not content:
        logger.warning("表格 %d 无内容文本，跳过提取规则应用", table_id)
        return

    try:
        key_info = await extract_key_info(
            content, extraction_rule_id, db,
            title=table_obj.name,
        )
        table_obj.extraction_rule_id = extraction_rule_id
        table_obj.key_info = key_info
        await db.commit()
        logger.info("提取规则 %d 已应用到表格 %d", extraction_rule_id, table_id)
    except Exception as e:
        logger.error("应用提取规则失败 (rule=%d, table=%d): %s", extraction_rule_id, table_id, e)


async def _apply_cleaning_after_import(
    db: AsyncSession, table_id: int, cleaning_rule_id: int
) -> None:
    """导入完成后自动应用清洗规则（内部辅助函数）。支持内置规则（负数 ID）和用户自定义规则。"""
    from app.models.cleaning_rule import CleaningRule
    from app.services.builtin_rules import get_builtin_cleaning_rule
    from app.services.structured_table_cleaner import apply_cleaning_rule

    if cleaning_rule_id < 0:
        # 内置规则：从内存中获取，构造一个临时对象
        builtin = get_builtin_cleaning_rule(cleaning_rule_id)
        if not builtin:
            logger.warning("内置清洗规则 %d 不存在，跳过清洗", cleaning_rule_id)
            return
        rule = CleaningRule(
            id=builtin["id"],
            name=builtin["name"],
            options=builtin["options"],
            field_hint=builtin.get("field_hint", ""),
        )
    else:
        rule = await db.get(CleaningRule, cleaning_rule_id)
        if not rule:
            logger.warning("清洗规则 %d 不存在，跳过清洗", cleaning_rule_id)
            return
    try:
        stats = await apply_cleaning_rule(db, table_id, rule)
        logger.info("清洗规则 %d 已应用到表格 %d: %s", cleaning_rule_id, table_id, stats)
    except Exception as e:
        logger.error("应用清洗规则失败 (rule=%d, table=%d): %s", cleaning_rule_id, table_id, e)


# ── 导入端点 ────────────────────────────────────────────────


@router.post("/import/bitable", response_model=StructuredTableOut, summary="从飞书多维表格导入")
async def import_bitable(
    body: ImportBitableRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from app.services.structured_table_import import import_from_bitable

    try:
        table = await import_from_bitable(
            db,
            current_user.feishu_open_id,
            body.app_token,
            body.table_id,
            user_access_token=current_user.feishu_access_token,
        )
        await db.commit()
        await db.refresh(table)
        # 如果指定了清洗规则，导入完成后自动应用
        if body.cleaning_rule_id:
            await _apply_cleaning_after_import(db, table.id, body.cleaning_rule_id)
            await db.refresh(table)
        # 如果指定了提取规则，导入完成后自动应用
        if body.extraction_rule_id:
            await _apply_extraction_after_import(db, table.id, body.extraction_rule_id)
            await db.refresh(table)
        # LLM 自动标签推荐
        await _auto_tag_structured_table(db, table, current_user.feishu_open_id)
        return table
    except Exception as e:
        logger.error("导入多维表格失败: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"导入失败: {e}")


@router.post("/import/spreadsheet", response_model=StructuredTableOut, summary="从飞书表格导入")
async def import_spreadsheet(
    body: ImportSpreadsheetRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from app.services.structured_table_import import import_from_spreadsheet

    try:
        table = await import_from_spreadsheet(
            db,
            current_user.feishu_open_id,
            body.spreadsheet_token,
            body.sheet_id,
            user_access_token=current_user.feishu_access_token,
        )
        await db.commit()
        await db.refresh(table)
        # 如果指定了清洗规则，导入完成后自动应用
        if body.cleaning_rule_id:
            await _apply_cleaning_after_import(db, table.id, body.cleaning_rule_id)
            await db.refresh(table)
        # 如果指定了提取规则，导入完成后自动应用
        if body.extraction_rule_id:
            await _apply_extraction_after_import(db, table.id, body.extraction_rule_id)
            await db.refresh(table)
        # LLM 自动标签推荐
        await _auto_tag_structured_table(db, table, current_user.feishu_open_id)
        return table
    except Exception as e:
        logger.error("导入飞书表格失败: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"导入失败: {e}")


@router.post("/import/upload", response_model=StructuredTableOut, summary="上传本地 CSV/Excel")
async def import_upload(
    file: UploadFile = File(...),
    cleaning_rule_id: int | None = Query(None, description="清洗规则 ID"),
    extraction_rule_id: int | None = Query(None, description="提取规则 ID"),
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    from app.services.structured_table_import import import_from_local_file

    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20MB 限制
        raise HTTPException(status_code=400, detail="文件大小不能超过 20MB")

    try:
        table = await import_from_local_file(
            db,
            current_user.feishu_open_id,
            file.filename,
            content,
        )
        await db.commit()
        await db.refresh(table)
        # 如果指定了清洗规则，导入完成后自动应用
        if cleaning_rule_id:
            await _apply_cleaning_after_import(db, table.id, cleaning_rule_id)
            await db.refresh(table)
        # 如果指定了提取规则，导入完成后自动应用
        if extraction_rule_id:
            await _apply_extraction_after_import(db, table.id, extraction_rule_id)
            await db.refresh(table)
        # LLM 自动标签推荐
        await _auto_tag_structured_table(db, table, current_user.feishu_open_id)
        return table
    except Exception as e:
        logger.error("导入本地文件失败: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"导入失败: {e}")


async def _auto_tag_structured_table(db: AsyncSession, table, owner_id: str) -> None:
    """对结构化表格调用 LLM 自动标签推荐（静默失败）。"""
    try:
        from app.services.llm import auto_tag_content
        content = f"{table.name} {table.description or ''} {getattr(table, 'summary', '') or ''}"
        tagged = await auto_tag_content(db, table.id, "structured_table", content.strip(), owner_id)
        if tagged:
            await db.commit()
            logger.info("自动标签推荐: structured_table_id=%d, 新增 %d 个标签", table.id, tagged)
    except Exception as e:
        logger.warning("自动标签推荐失败 (structured_table_id=%d): %s", table.id, e)


# ── URL 解析 & 导入 ──────────────────────────────────────────────


import re


def _parse_feishu_url(url: str) -> dict:
    """解析飞书链接，提取类型和 token。

    支持的 URL 格式：
    - 多维表格: https://xxx.feishu.cn/base/{app_token}?table={table_id}
    - Wiki 内嵌多维表格: https://xxx.feishu.cn/wiki/{node_token}?table={table_id}
    - 飞书表格: https://xxx.feishu.cn/sheets/{token}
    - Wiki 内嵌飞书表格: https://xxx.feishu.cn/wiki/{node_token} (obj_type=sheet)
    """
    url = url.strip()

    # 多维表格直链: /base/{app_token}
    m = re.search(r'/base/([A-Za-z0-9_-]+)', url)
    if m:
        app_token = m.group(1)
        # 尝试提取 table_id
        tm = re.search(r'[?&]table=([A-Za-z0-9_-]+)', url)
        table_id = tm.group(1) if tm else None
        return {"type": "bitable", "token": app_token, "table_id": table_id}

    # 飞书表格直链: /sheets/{token}
    m = re.search(r'/sheets/([A-Za-z0-9_-]+)', url)
    if m:
        return {"type": "spreadsheet", "token": m.group(1), "table_id": None}

    # Wiki 链接: /wiki/{node_token}
    m = re.search(r'/wiki/([A-Za-z0-9_-]+)', url)
    if m:
        node_token = m.group(1)
        tm = re.search(r'[?&]table=([A-Za-z0-9_-]+)', url)
        table_id = tm.group(1) if tm else None
        return {"type": "wiki", "token": node_token, "table_id": table_id}

    raise ValueError(
        "无法识别的链接格式，请粘贴飞书多维表格或飞书表格的链接"
    )


@router.post("/parse-url", response_model=URLParseResult, summary="解析飞书链接")
async def parse_feishu_url(
    body: ImportFromURLRequest,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """解析飞书链接，返回类型、token 和子表/工作表列表。"""
    from app.services.feishu import feishu_client, FeishuAPIError

    try:
        parsed = _parse_feishu_url(body.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    user_token = current_user.feishu_access_token

    # 如果是 wiki 链接，先解析出实际的 obj_token 和 obj_type
    # 注意：用 tenant_access_token 解析 Wiki 节点（避免用户 token 缺少 wiki 权限）
    if parsed["type"] == "wiki":
        try:
            # 先尝试用户 token，失败再用应用 token
            try:
                node_info = await feishu_client.get_wiki_node_info(
                    parsed["token"], user_access_token=user_token,
                )
            except Exception:
                logger.info("用户 token 解析 Wiki 节点失败，改用应用 token")
                node_info = await feishu_client.get_wiki_node_info(
                    parsed["token"],
                )
            logger.info("Wiki 节点解析结果: %s", node_info)
            obj_type = node_info.get("obj_type", "")
            obj_token = node_info.get("obj_token", "")
            if obj_type == "bitable":
                parsed = {"type": "bitable", "token": obj_token, "table_id": parsed["table_id"]}
            elif obj_type == "sheet":
                parsed = {"type": "spreadsheet", "token": obj_token, "table_id": None}
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"该 Wiki 页面类型是 {obj_type}，不是多维表格或飞书表格",
                )
        except HTTPException:
            raise
        except Exception as e:
            logger.error("解析 Wiki 链接失败: %s", e, exc_info=True)
            raise HTTPException(status_code=400, detail=f"解析 Wiki 链接失败: {e}")

    try:
        if parsed["type"] == "bitable":
            # 获取多维表格下的数据表列表
            tables_raw = await feishu_client.get_bitable_tables(
                parsed["token"], user_access_token=user_token,
            )
            tables = [{"table_id": t.get("table_id", ""), "name": t.get("name", "")} for t in tables_raw]

            return URLParseResult(
                source_type="bitable",
                app_token=parsed["token"],
                table_id=parsed.get("table_id"),
                tables=tables,
            )

        elif parsed["type"] == "spreadsheet":
            # 获取飞书表格下的工作表列表
            sheets_raw = await feishu_client.get_spreadsheet_sheets(
                parsed["token"], user_access_token=user_token,
            )
            sheets = [{"sheet_id": s.get("sheet_id", ""), "title": s.get("title", "")} for s in sheets_raw]

            return URLParseResult(
                source_type="spreadsheet",
                app_token=parsed["token"],
                sheets=sheets,
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("获取子表/工作表列表失败: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"获取表格信息失败: {e}")

    raise HTTPException(status_code=400, detail="无法识别的链接类型")


@router.post("/import/url", response_model=StructuredTableOut, summary="通过飞书链接导入")
async def import_from_url(
    body: ImportFromURLRequest,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """通过飞书链接直接导入（自动解析链接类型）。"""
    from app.services.feishu import feishu_client, FeishuAPIError
    from app.services.structured_table_import import import_from_bitable, import_from_spreadsheet

    try:
        parsed = _parse_feishu_url(body.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    user_token = current_user.feishu_access_token

    # Wiki 链接解析（先尝试用户 token，失败再用应用 token）
    if parsed["type"] == "wiki":
        try:
            try:
                node_info = await feishu_client.get_wiki_node_info(
                    parsed["token"], user_access_token=user_token,
                )
            except Exception:
                node_info = await feishu_client.get_wiki_node_info(
                    parsed["token"],
                )
            obj_type = node_info.get("obj_type", "")
            obj_token = node_info.get("obj_token", "")
            if obj_type == "bitable":
                parsed = {"type": "bitable", "token": obj_token, "table_id": parsed["table_id"]}
            elif obj_type == "sheet":
                parsed = {"type": "spreadsheet", "token": obj_token, "table_id": None}
            else:
                raise HTTPException(status_code=400, detail=f"该页面类型 {obj_type} 不支持导入")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"解析 Wiki 链接失败: {e}")

    try:
        if parsed["type"] == "bitable":
            table_id = parsed.get("table_id")
            if not table_id:
                # URL 里没有 table_id，取第一个子表
                tables_raw = await feishu_client.get_bitable_tables(
                    parsed["token"], user_access_token=user_token,
                )
                if not tables_raw:
                    raise HTTPException(status_code=400, detail="该多维表格下没有数据表")
                table_id = tables_raw[0].get("table_id", "")

            table = await import_from_bitable(
                db, current_user.feishu_open_id,
                parsed["token"], table_id,
                user_access_token=user_token,
            )
            await db.commit()
            await db.refresh(table)
            if body.cleaning_rule_id:
                await _apply_cleaning_after_import(db, table.id, body.cleaning_rule_id)
                await db.refresh(table)
            if body.extraction_rule_id:
                await _apply_extraction_after_import(db, table.id, body.extraction_rule_id)
                await db.refresh(table)
            return table

        elif parsed["type"] == "spreadsheet":
            # 取第一个工作表
            sheets_raw = await feishu_client.get_spreadsheet_sheets(
                parsed["token"], user_access_token=user_token,
            )
            if not sheets_raw:
                raise HTTPException(status_code=400, detail="该飞书表格下没有工作表")
            sheet_id = sheets_raw[0].get("sheet_id", "")

            table = await import_from_spreadsheet(
                db, current_user.feishu_open_id,
                parsed["token"], sheet_id,
                user_access_token=user_token,
            )
            await db.commit()
            await db.refresh(table)
            if body.cleaning_rule_id:
                await _apply_cleaning_after_import(db, table.id, body.cleaning_rule_id)
                await db.refresh(table)
            if body.extraction_rule_id:
                await _apply_extraction_after_import(db, table.id, body.extraction_rule_id)
                await db.refresh(table)
            return table

    except HTTPException:
        raise
    except Exception as e:
        logger.error("通过链接导入失败: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"导入失败: {e}")

    raise HTTPException(status_code=400, detail="无法识别的链接类型")


# ── 发现飞书表格（必须在 /{table_id} 之前注册） ─────────────────


@router.get("/discover-spreadsheets", summary="发现用户的飞书表格")
async def discover_spreadsheets(
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """列出用户有权限访问的飞书表格（sheet 类型），包含云空间 + 知识空间。"""
    from app.services.feishu import feishu_client

    seen_tokens: set[str] = set()
    all_files: list[dict] = []

    # 1. 云空间中的飞书表格
    try:
        drive_files = await feishu_client.list_drive_spreadsheets(
            user_access_token=current_user.feishu_access_token,
        )
        for f in drive_files:
            t = f.get("token", "")
            if t and t not in seen_tokens:
                seen_tokens.add(t)
                all_files.append({"token": t, "name": f.get("name", "")})
    except Exception as e:
        logger.warning("获取云空间飞书表格失败: %s", e)

    # 2. 知识空间（Wiki）中的飞书表格
    try:
        wiki_nodes = await feishu_client.list_wiki_nodes_by_type(
            {"sheet"},
            user_access_token=current_user.feishu_access_token,
        )
        for node in wiki_nodes:
            t = node.get("obj_token", "")
            if t and t not in seen_tokens:
                seen_tokens.add(t)
                space_name = node.get("space_name", "")
                title = node.get("title", "未命名")
                name = f"[{space_name}] {title}" if space_name else title
                all_files.append({"token": t, "name": name})
    except Exception as e:
        logger.warning("获取知识空间飞书表格失败: %s", e)

    if not all_files:
        raise HTTPException(status_code=400, detail="未找到任何飞书表格，请检查飞书权限")

    return {"files": all_files}


@router.get("/discover-sheets/{spreadsheet_token}", summary="获取飞书表格下的工作表列表")
async def discover_sheets(
    spreadsheet_token: str,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """获取指定飞书表格下的所有工作表。"""
    from app.services.feishu import feishu_client

    try:
        sheets = await feishu_client.get_spreadsheet_sheets(
            spreadsheet_token,
            user_access_token=current_user.feishu_access_token,
        )
        return {"sheets": [{"sheet_id": s.get("sheet_id", ""), "title": s.get("title", "")} for s in sheets]}
    except Exception as e:
        logger.error("获取工作表列表失败: %s", e)
        raise HTTPException(status_code=400, detail=f"获取列表失败: {e}")


# ── 列表 & 详情 ──────────────────────────────────────────────


@router.get("", response_model=StructuredTableListResponse, summary="表格列表")
async def list_tables(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str = Query("", max_length=200),
    source_type: str = Query("", max_length=32),
    table_category: str | None = Query(None),
    asset_owner_name: str | None = Query(None),
    tag_ids: list[int] = Query(default=[]),
    extraction_rule_id: int | None = Query(None, description="按提取规则ID筛选，-999表示无规则"),
    date_field: str | None = Query(None, description="时间筛选字段: synced_at, created_at, updated_at"),
    date_from: datetime | None = Query(None, description="时间范围开始"),
    date_to: datetime | None = Query(None, description="时间范围结束"),
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    visible_ids = await get_visible_owner_ids(current_user, db, request)

    conditions = []
    if visible_ids is not None:
        conditions.append(StructuredTable.owner_id.in_(visible_ids))
    if search:
        conditions.append(StructuredTable.name.ilike(f"%{search}%"))
    if source_type:
        conditions.append(StructuredTable.source_type == source_type)
    if table_category:
        conditions.append(StructuredTable.table_category == table_category)
    if extraction_rule_id is not None:
        if extraction_rule_id == -999:
            conditions.append(StructuredTable.extraction_rule_id.is_(None))
        else:
            conditions.append(StructuredTable.extraction_rule_id == extraction_rule_id)

    if tag_ids:
        subq = select(ContentTag.content_id).where(
            ContentTag.content_type == "structured_table",
            ContentTag.tag_id.in_(tag_ids),
        )
        conditions.append(StructuredTable.id.in_(subq))

    # 时间范围筛选
    _date_field_map = {
        "synced_at": StructuredTable.synced_at,
        "created_at": StructuredTable.created_at,
        "updated_at": StructuredTable.updated_at,
    }
    if date_field and date_field in _date_field_map:
        col = _date_field_map[date_field]
        if date_from:
            conditions.append(col >= date_from)
        if date_to:
            conditions.append(col <= date_to)

    # 总数
    count_q = select(func.count()).select_from(StructuredTable)
    if conditions:
        count_q = count_q.where(and_(*conditions))
    total = (await db.execute(count_q)).scalar() or 0

    # asset_owner_name 过滤需要 join User 表
    if asset_owner_name:
        user_subq = select(User.feishu_open_id).where(User.name.ilike(f"%{asset_owner_name}%"))
        conditions.append(StructuredTable.owner_id.in_(user_subq))
        # 重新计算总数
        count_q = select(func.count()).select_from(StructuredTable).where(and_(*conditions))
        total = (await db.execute(count_q)).scalar() or 0

    # 分页（LEFT JOIN User 获取 asset_owner_name）
    q = select(StructuredTable, User.name.label("asset_owner_name")).outerjoin(User, StructuredTable.owner_id == User.feishu_open_id)
    if conditions:
        q = q.where(and_(*conditions))
    q = (q
        .order_by(func.coalesce(StructuredTable.synced_at, StructuredTable.created_at).desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(q)
    rows = result.all()

    # 统计每个 (source_app_token, source_table_id) 被多少人归档（按上传人去重）
    source_keys = [
        (r.StructuredTable.source_app_token, r.StructuredTable.source_table_id)
        for r in rows
        if r.StructuredTable.source_app_token and r.StructuredTable.source_table_id
    ]
    import_count_map: dict[tuple[str, str], int] = {}
    if source_keys:
        unique_app_tokens = list({k[0] for k in source_keys})
        count_rows = (await db.execute(
            select(
                StructuredTable.source_app_token,
                StructuredTable.source_table_id,
                func.count(StructuredTable.owner_id.distinct()).label("cnt"),
            )
            .where(
                StructuredTable.source_app_token.in_(unique_app_tokens),
                StructuredTable.source_table_id.isnot(None),
            )
            .group_by(StructuredTable.source_app_token, StructuredTable.source_table_id)
        )).all()
        import_count_map = {
            (row.source_app_token, row.source_table_id): row.cnt
            for row in count_rows
        }

    # 批量查询当前页所有表格的标签
    from app.schemas.document import ContentTagBrief
    table_ids = [r.StructuredTable.id for r in rows]
    tags_map: dict[int, list[ContentTagBrief]] = {}
    if table_ids:
        tag_rows = (await db.execute(
            select(ContentTag.content_id, ContentTag.id.label("ct_id"), ContentTag.tag_id,
                   TagDefinition.name, TagDefinition.color)
            .join(TagDefinition, ContentTag.tag_id == TagDefinition.id)
            .where(ContentTag.content_type == "structured_table",
                   ContentTag.content_id.in_(table_ids))
        )).all()
        for row in tag_rows:
            tags_map.setdefault(row.content_id, []).append(
                ContentTagBrief(id=row.ct_id, tag_id=row.tag_id,
                                tag_name=row.name, tag_color=row.color))

    # 批量查询清洗规则名称
    from app.models.cleaning_rule import CleaningRule
    rule_ids = list({r.StructuredTable.cleaning_rule_id for r in rows if r.StructuredTable.cleaning_rule_id})
    cleaning_rule_map: dict[int, str] = {}
    if rule_ids:
        rule_rows = (await db.execute(
            select(CleaningRule.id, CleaningRule.name).where(CleaningRule.id.in_(rule_ids))
        )).all()
        cleaning_rule_map = {row.id: row.name for row in rule_rows}

    items = []
    for r in rows:
        tbl = r.StructuredTable
        out = StructuredTableOut.model_validate(tbl)
        # 飞书资产：优先使用 asset_owner_name 字段（真实所有者），其次 extra_fields，最后 User 表 join（导入者）
        feishu_owner = tbl.asset_owner_name or (tbl.extra_fields or {}).get("_feishu_owner_name", "")
        display_owner = feishu_owner or r.asset_owner_name
        updates: dict = {"asset_owner_name": display_owner, "uploader_name": r.asset_owner_name}
        key = (tbl.source_app_token, tbl.source_table_id)
        if key[0] and key[1] and key in import_count_map:
            updates["import_count"] = import_count_map[key]
        if tbl.cleaning_rule_id and tbl.cleaning_rule_id in cleaning_rule_map:
            updates["cleaning_rule_name"] = cleaning_rule_map[tbl.cleaning_rule_id]
        if tbl.id in tags_map:
            updates["tags"] = tags_map[tbl.id]
        out = out.model_copy(update=updates)
        items.append(out)

    return StructuredTableListResponse(items=items, total=total)


@router.get("/categories", summary="获取表格分类列表")
async def list_table_categories(
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """返回可见范围内所有表格的 table_category 去重列表（动态，由 AI 识别）。"""
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    stmt = select(StructuredTable.table_category).where(
        StructuredTable.table_category.isnot(None),
    )
    if visible_ids is not None:
        stmt = stmt.where(StructuredTable.owner_id.in_(visible_ids))
    stmt = stmt.distinct().order_by(StructuredTable.table_category)
    rows = (await db.execute(stmt)).scalars().all()
    return {"categories": rows}


@router.get("/search", response_model=SearchResponse, summary="穿透搜索")
async def search_rows(
    q: str = Query(..., min_length=1, max_length=200, description="搜索关键词"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """穿透搜索：在所有表格的行数据中搜索关键词。"""
    owner_id = current_user.feishu_open_id
    keyword = q.strip()

    # 联表查询：行数据 ilike 匹配 + 表级权限过滤
    count_q = (
        select(func.count())
        .select_from(StructuredTableRow)
        .join(StructuredTable, StructuredTableRow.table_id == StructuredTable.id)
        .where(and_(
            StructuredTable.owner_id == owner_id,
            StructuredTableRow.row_text.ilike(f"%{keyword}%"),
        ))
    )
    total = (await db.execute(count_q)).scalar() or 0

    q_rows = (
        select(StructuredTableRow, StructuredTable.name, StructuredTable.schema_info)
        .join(StructuredTable, StructuredTableRow.table_id == StructuredTable.id)
        .where(and_(
            StructuredTable.owner_id == owner_id,
            StructuredTableRow.row_text.ilike(f"%{keyword}%"),
        ))
        .order_by(StructuredTableRow.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(q_rows)
    rows = result.all()

    results: list[SearchResultItem] = []
    keyword_lower = keyword.lower()
    for row_obj, table_name, schema_info in rows:
        # 构建 field_id -> field_name 映射，将原始字段ID翻译为中文
        field_map: dict[str, str] = {}
        if schema_info:
            for s in schema_info:
                fid = s.get("field_id", "")
                fname = s.get("field_name", "")
                if fid and fname:
                    field_map[fid] = fname

        # 翻译 row_data 的 key 为中文字段名
        raw_data = row_obj.row_data or {}
        row_data = {field_map.get(k, k): v for k, v in raw_data.items()}

        # 找出匹配的字段（使用翻译后的名称）
        matched_fields = []
        for field_name, value in row_data.items():
            if value and keyword_lower in str(value).lower():
                matched_fields.append(field_name)

        results.append(SearchResultItem(
            table_id=row_obj.table_id,
            table_name=table_name,
            row_id=row_obj.id,
            row_index=row_obj.row_index,
            row_data=row_data,
            matched_fields=matched_fields,
        ))

    return SearchResponse(keyword=keyword, total=total, results=results)


@router.get("/{table_id}/archivers", summary="表格归档人列表")
async def get_table_archivers(
    table_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """查询该表格的所有归档人（按 source_app_token + source_table_id 聚合）。"""
    row = (await db.execute(
        select(StructuredTable.source_app_token, StructuredTable.source_table_id)
        .where(StructuredTable.id == table_id)
    )).first()
    if row is None:
        raise HTTPException(status_code=404, detail="表格不存在")
    if not row.source_app_token or not row.source_table_id:
        return {"archivers": []}

    rows = (await db.execute(
        select(
            StructuredTable.owner_id,
            func.coalesce(User.name, StructuredTable.asset_owner_name).label("name"),
            User.avatar_url,
            func.min(StructuredTable.created_at).label("archived_at"),
        )
        .outerjoin(User, StructuredTable.owner_id == User.feishu_open_id)
        .where(
            StructuredTable.source_app_token == row.source_app_token,
            StructuredTable.source_table_id == row.source_table_id,
        )
        .group_by(StructuredTable.owner_id, User.name, StructuredTable.asset_owner_name, User.avatar_url)
    )).all()

    return {"archivers": [
        {
            "name": r.name or "未知用户",
            "avatar_url": r.avatar_url,
            "archived_at": r.archived_at.isoformat() if r.archived_at else None,
        }
        for r in rows
    ]}


@router.get("/{table_id}", response_model=StructuredTableDetail, summary="表格详情")
async def get_table(
    request: Request,
    table_id: int,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        stmt = stmt.where(StructuredTable.owner_id.in_(visible_ids))
    result = await db.execute(stmt)
    table = result.scalar_one_or_none()
    if not table:
        raise HTTPException(status_code=404, detail="表格不存在或无权访问")
    out = StructuredTableDetail.model_validate(table)
    # 从 extra_fields 提取 sheet_names
    if table.extra_fields and "sheet_names" in table.extra_fields:
        out = out.model_copy(update={"sheet_names": table.extra_fields["sheet_names"]})
    if table.cleaning_rule_id:
        from app.models.cleaning_rule import CleaningRule
        rule = await db.get(CleaningRule, table.cleaning_rule_id)
        if rule:
            out = out.model_copy(update={"cleaning_rule_name": rule.name})
    return out


@router.get("/{table_id}/rows", response_model=StructuredTableRowListResponse, summary="行数据预览")
async def get_table_rows(
    request: Request,
    table_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str = Query("", max_length=200),
    sheet_name: str = Query("", max_length=256),
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    # 校验归属
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    chk_stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        chk_stmt = chk_stmt.where(StructuredTable.owner_id.in_(visible_ids))
    table_result = await db.execute(chk_stmt)
    if not table_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="表格不存在或无权访问")

    conditions = [StructuredTableRow.table_id == table_id]
    if sheet_name:
        conditions.append(StructuredTableRow.sheet_name == sheet_name)
    if search:
        conditions.append(StructuredTableRow.row_text.ilike(f"%{search}%"))

    count_q = select(func.count()).select_from(StructuredTableRow).where(and_(*conditions))
    total = (await db.execute(count_q)).scalar() or 0

    q = (
        select(StructuredTableRow)
        .where(and_(*conditions))
        .order_by(StructuredTableRow.row_index)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(q)
    items = result.scalars().all()

    return StructuredTableRowListResponse(items=items, total=total)


# ── 同步 & 删除 ──────────────────────────────────────────────


@router.post("/{table_id}/sync", summary="重新同步飞书源")
async def sync_table(
    request: Request,
    table_id: int,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    from app.services.structured_table_import import sync_table as do_sync

    # 校验归属
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    chk_stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        chk_stmt = chk_stmt.where(StructuredTable.owner_id.in_(visible_ids))
    table_result = await db.execute(chk_stmt)
    if not table_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="表格不存在或无权访问")

    try:
        result = await do_sync(db, table_id, user_access_token=current_user.feishu_access_token)
        # 如果表格已绑定清洗规则，同步后自动重新应用
        table_obj = await db.get(StructuredTable, table_id)
        if table_obj and table_obj.cleaning_rule_id:
            await _apply_cleaning_after_import(db, table_id, table_obj.cleaning_rule_id)
        # 如果表格已绑定提取规则，同步后自动重新应用
        if table_obj and table_obj.extraction_rule_id:
            await _apply_extraction_after_import(db, table_id, table_obj.extraction_rule_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("同步表格失败: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"同步失败: {e}")


@router.patch("/{table_id}/extraction-rule", summary="绑定或修改提取规则")
async def update_extraction_rule(
    request: Request,
    table_id: int,
    body: dict,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """绑定、修改或解除表格的提取规则。body: { extraction_rule_id: int | null }"""
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        stmt = stmt.where(StructuredTable.owner_id.in_(visible_ids))
    result = await db.execute(stmt)
    table = result.scalar_one_or_none()
    if not table:
        raise HTTPException(status_code=404, detail="表格不存在或无权操作")

    new_rule_id = body.get("extraction_rule_id")
    if new_rule_id is None:
        # 解绑规则
        table.extraction_rule_id = None
        table.key_info = None
        await db.commit()
        return {"message": "已解除提取规则", "extraction_rule_id": None}

    # 绑定并立即应用提取规则
    await _apply_extraction_after_import(db, table_id, new_rule_id)
    await db.refresh(table)
    return {
        "message": "提取规则已应用",
        "extraction_rule_id": table.extraction_rule_id,
        "key_info": table.key_info,
    }


async def _restore_original_data(db: AsyncSession, table: StructuredTable) -> None:
    """从原始来源重新导入数据，恢复到未清洗状态。

    对本地文件：直接重新解析原始文件并替换当前表的行数据（不创建新表）。
    对飞书来源：重新同步后将新表数据迁移回原表。
    """
    from app.services.structured_table_import import (
        import_from_bitable, import_from_local_file, import_from_spreadsheet,
    )

    if table.source_type == "local" and table.file_path:
        import os
        if not os.path.exists(table.file_path):
            logger.warning("原始文件不存在: %s，无法恢复", table.file_path)
            return

        # 重新解析原始文件，将行数据写回当前表（不创建新表）
        from app.services.structured_table_import import (
            _parse_excel_xlsx, _parse_excel_xls, _parse_csv,
            _deduplicate_headers, build_row_text,
        )
        import io

        with open(table.file_path, "rb") as f:
            file_content = f.read()
        file_name = table.file_name or os.path.basename(table.file_path)
        lower_name = file_name.lower()

        if lower_name.endswith(".csv"):
            parsed = _parse_csv(file_content)
        elif lower_name.endswith(".tsv"):
            parsed = _parse_csv(file_content, delimiter="\t")
        elif lower_name.endswith(".xlsx"):
            parsed = _parse_excel_xlsx(file_content)
        elif lower_name.endswith(".xls"):
            parsed = _parse_excel_xls(file_content)
        else:
            logger.warning("不支持的文件格式: %s，无法恢复", file_name)
            return

        # 统一为 sheets 格式
        if isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], tuple) and len(parsed[0]) == 3:
            sheets = parsed
        elif isinstance(parsed, tuple) and len(parsed) == 2:
            sheets = [(None, parsed[0], parsed[1])]
        else:
            logger.warning("文件解析失败，无法恢复: %s", file_name)
            return

        # 删除旧行
        from sqlalchemy import delete as sa_delete
        await db.execute(
            sa_delete(StructuredTableRow).where(StructuredTableRow.table_id == table.id)
        )

        # 插入新行
        is_multi_sheet = len(sheets) > 1
        total_rows = 0
        text_rows = []
        first_headers = None
        for sheet_name, headers, data_rows in sheets:
            unique_headers = _deduplicate_headers(headers)
            if first_headers is None:
                first_headers = unique_headers
            for idx, row in enumerate(data_rows):
                row_data = {}
                for col_idx, header in enumerate(unique_headers):
                    val = row[col_idx] if col_idx < len(row) else None
                    row_data[header] = str(val) if val is not None else None
                db.add(StructuredTableRow(
                    table_id=table.id,
                    sheet_name=sheet_name,
                    row_index=total_rows,
                    row_data=row_data,
                    row_text=build_row_text(row_data),
                ))
                total_rows += 1
                if len(text_rows) < 50:
                    text_rows.append(" | ".join(f"{k}: {v}" for k, v in row_data.items() if v))

        # 更新表级元数据
        if is_multi_sheet:
            table.schema_info = {
                "__sheets__": {
                    name: [
                        {"field_id": f"col_{i}", "field_name": h, "field_type": "text"}
                        for i, h in enumerate(_deduplicate_headers(headers))
                    ]
                    for name, headers, _ in sheets
                }
            }
        else:
            table.schema_info = [
                {"field_id": f"col_{i}", "field_name": h, "field_type": "text"}
                for i, h in enumerate(first_headers or [])
            ]
        table.row_count = total_rows
        table.column_count = len(first_headers or [])
        table.cleaning_rule_id = None
        table.content_text = "\n".join(text_rows)
        await db.flush()
        logger.info("已从原始文件恢复表格 %d 的数据 (%d 行)", table.id, total_rows)

    elif table.source_type == "bitable" and table.source_app_token:
        from app.worker.tasks import _resolve_user_token
        token = await _resolve_user_token(table.owner_id, db)
        if token:
            await import_from_bitable(
                db, table.owner_id, table.source_app_token,
                table.source_table_id, user_access_token=token,
            )
            logger.info("已从飞书重新同步表格 %d 的数据", table.id)
    elif table.source_type == "spreadsheet" and table.source_app_token:
        from app.worker.tasks import _resolve_user_token
        token = await _resolve_user_token(table.owner_id, db)
        if token:
            await import_from_spreadsheet(
                db, table.owner_id, table.source_app_token,
                table.source_table_id, user_access_token=token,
            )
            logger.info("已从飞书重新同步表格 %d 的数据", table.id)


@router.patch("/{table_id}/cleaning-rule", summary="绑定或修改清洗规则")
async def update_cleaning_rule(
    request: Request,
    table_id: int,
    body: dict,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """绑定、修改或解除表格的清洗规则。body: { cleaning_rule_id: int | null }"""
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        stmt = stmt.where(StructuredTable.owner_id.in_(visible_ids))
    result = await db.execute(stmt)
    table = result.scalar_one_or_none()
    if not table:
        raise HTTPException(status_code=404, detail="表格不存在或无权操作")

    new_rule_id = body.get("cleaning_rule_id")

    # 无论是解绑还是切换规则，都先从原始来源恢复数据
    if table.cleaning_rule_id is not None:
        await _restore_original_data(db, table)
        # 重新加载 table 对象（import 会更新它）
        await db.refresh(table)

    if new_rule_id is None:
        # 解绑规则，数据已恢复
        table.cleaning_rule_id = None
        await db.commit()
        return {"message": "已解除清洗规则，数据已恢复", "cleaning_rule_id": None}

    # 绑定并在原始数据上应用新清洗规则
    await _apply_cleaning_after_import(db, table.id, new_rule_id)
    await db.refresh(table)
    return {
        "message": "清洗规则已应用",
        "cleaning_rule_id": table.cleaning_rule_id,
        "row_count": table.row_count,
        "column_count": table.column_count,
    }


@router.delete("/{table_id}", summary="删除表格")
async def delete_table(
    request: Request,
    table_id: int,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        stmt = stmt.where(StructuredTable.owner_id.in_(visible_ids))
    result = await db.execute(stmt)
    table = result.scalar_one_or_none()
    if not table:
        raise HTTPException(status_code=404, detail="表格不存在或无权删除")

    # 清理原始文件
    if table.file_path:
        import os
        try:
            if os.path.exists(table.file_path):
                os.remove(table.file_path)
        except OSError:
            logger.warning("删除原始文件失败: %s", table.file_path)

    # 清理关联的标签和知识图谱实体链接
    await db.execute(
        delete(ContentTag).where(
            ContentTag.content_type == "structured_table",
            ContentTag.content_id == table_id,
        )
    )
    await db.execute(
        delete(ContentEntityLink).where(
            ContentEntityLink.content_type == "structured_table",
            ContentEntityLink.content_id == table_id,
        )
    )

    # 删除关联的飞书数据源和同步状态（这样用户可以重新添加）
    if table.source_app_token and table.source_table_id:
        await db.execute(
            delete(ETLDataSource).where(
                ETLDataSource.app_token == table.source_app_token,
                ETLDataSource.table_id == table.source_table_id,
            )
        )
        await db.execute(
            delete(ETLSyncState).where(
                ETLSyncState.source_app_token == table.source_app_token,
                ETLSyncState.source_table_id == table.source_table_id,
            )
        )

    # 级联删除行（靠 FK ON DELETE CASCADE），但显式删除更安全
    await db.execute(
        delete(StructuredTableRow).where(StructuredTableRow.table_id == table_id)
    )
    await db.delete(table)
    await db.commit()
    return {"detail": "已删除"}


@router.post("/batch-delete", summary="批量删除")
async def batch_delete_tables(
    request: Request,
    body: BatchDeleteRequest,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    if not body.ids:
        return {"deleted": 0}

    # 先根据 visible_ids 筛出允许删除的表格 ID
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    chk_stmt = select(StructuredTable.id).where(StructuredTable.id.in_(body.ids))
    if visible_ids is not None:
        chk_stmt = chk_stmt.where(StructuredTable.owner_id.in_(visible_ids))
    allowed = [row[0] for row in (await db.execute(chk_stmt)).fetchall()]
    if not allowed:
        return {"deleted": 0}

    # 清理关联的标签和知识图谱实体链接
    await db.execute(
        delete(ContentTag).where(
            ContentTag.content_type == "structured_table",
            ContentTag.content_id.in_(allowed),
        )
    )
    await db.execute(
        delete(ContentEntityLink).where(
            ContentEntityLink.content_type == "structured_table",
            ContentEntityLink.content_id.in_(allowed),
        )
    )

    # 删除关联的飞书数据源和同步状态
    source_rows = (await db.execute(
        select(StructuredTable.source_app_token, StructuredTable.source_table_id).where(
            StructuredTable.id.in_(allowed),
            StructuredTable.source_app_token.isnot(None),
            StructuredTable.source_table_id.isnot(None),
        )
    )).fetchall()
    for app_token, tbl_id in source_rows:
        await db.execute(
            delete(ETLDataSource).where(
                ETLDataSource.app_token == app_token,
                ETLDataSource.table_id == tbl_id,
            )
        )
        await db.execute(
            delete(ETLSyncState).where(
                ETLSyncState.source_app_token == app_token,
                ETLSyncState.source_table_id == tbl_id,
            )
        )

    # 先删行
    await db.execute(
        delete(StructuredTableRow).where(
            StructuredTableRow.table_id.in_(allowed)
        )
    )
    # 再删表
    result = await db.execute(
        delete(StructuredTable).where(StructuredTable.id.in_(allowed))
    )
    await db.commit()
    return {"deleted": result.rowcount}


# ── 导出 XLSX ────────────────────────────────────────────────


from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@router.get("/{table_id}/export", summary="导出为 XLSX 文件")
async def export_table_xlsx(
    request: Request,
    table_id: int,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """导出结构化表格为 Excel 文件。"""
    # 校验归属
    visible_ids = await get_visible_owner_ids(current_user, db, request)
    chk_stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        chk_stmt = chk_stmt.where(StructuredTable.owner_id.in_(visible_ids))
    result = await db.execute(chk_stmt)
    table = result.scalar_one_or_none()
    if not table:
        raise HTTPException(status_code=404, detail="表格不存在或无权访问")

    # 获取所有行数据
    rows_result = await db.execute(
        select(StructuredTableRow)
        .where(StructuredTableRow.table_id == table_id)
        .order_by(StructuredTableRow.sheet_name.nullsfirst(), StructuredTableRow.row_index)
    )
    rows = rows_result.scalars().all()

    # 检查是否多 sheet
    sheet_names = table.extra_fields.get("sheet_names") if table.extra_fields else None

    # 创建 Excel 工作簿
    wb = Workbook()
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def _write_sheet(ws, sheet_rows):
        columns: list[str] = []
        if sheet_rows and sheet_rows[0].row_data:
            columns = list(sheet_rows[0].row_data.keys())
        if not columns:
            columns = ["序号"]
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        for row_idx, row in enumerate(sheet_rows, 2):
            if row.row_data:
                for col_idx, col_name in enumerate(columns, 1):
                    value = row.row_data.get(col_name, "")
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in sheet_rows:
                if row.row_data:
                    cell_value = row.row_data.get(col_name, "")
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    if sheet_names and len(sheet_names) > 1:
        # 多 sheet 导出
        wb.remove(wb.active)
        for sn in sheet_names:
            ws = wb.create_sheet(title=sn[:31])  # Excel sheet 名最长 31 字符
            sheet_rows = [r for r in rows if r.sheet_name == sn]
            _write_sheet(ws, sheet_rows)
    else:
        ws = wb.active
        ws.title = "数据"
        _write_sheet(ws, rows)

    # 写入内存缓冲区
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 生成文件名（中文需要 URL 编码）
    from urllib.parse import quote
    filename = f"{table.name}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/{table_id}/download-original", summary="下载原始上传文件")
async def download_original_file(
    request: Request,
    table_id: int,
    current_user: Annotated[User, Depends(get_current_user)] = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """下载本地上传表格的原始文件。"""
    import os
    from fastapi.responses import FileResponse

    visible_ids = await get_visible_owner_ids(current_user, db, request)
    stmt = select(StructuredTable).where(StructuredTable.id == table_id)
    if visible_ids is not None:
        stmt = stmt.where(StructuredTable.owner_id.in_(visible_ids))
    result = await db.execute(stmt)
    table = result.scalar_one_or_none()
    if not table:
        raise HTTPException(status_code=404, detail="表格不存在或无权访问")
    if not table.file_path or not os.path.exists(table.file_path):
        raise HTTPException(status_code=404, detail="原始文件不存在（仅本地上传的表格保留原始文件）")

    return FileResponse(
        path=table.file_path,
        filename=table.file_name or f"original_{table_id}",
        media_type="application/octet-stream",
    )
