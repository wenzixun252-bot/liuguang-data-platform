"""提取规则 CRUD API。"""

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from sqlalchemy import String, func, select, union_all, literal_column, literal
from sqlalchemy.sql.expression import cast
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db, get_visible_owner_ids
from app.models.extraction_rule import ExtractionRule
from app.models.document import Document
from app.models.communication import Communication
from app.models.structured_table import StructuredTable
from app.models.user import User
from app.schemas.extraction_rule import (
    ExtractionRuleCreate,
    ExtractionRuleOut,
    ExtractionRuleUpdate,
    ExtractionDataItem,
    ExtractionDataResponse,
)
from app.services.extraction_templates import SECTOR_LABELS, SECTOR_TEMPLATES, get_template_fields

router = APIRouter(prefix="/api/extraction-rules", tags=["extraction-rules"])


@router.get("", response_model=list[ExtractionRuleOut])
async def list_rules(
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    result = await db.execute(
        select(ExtractionRule)
        .where(ExtractionRule.owner_id == user.feishu_open_id)
        .order_by(ExtractionRule.updated_at.desc())
    )
    return result.scalars().all()


@router.post("", response_model=ExtractionRuleOut)
async def create_rule(
    body: ExtractionRuleCreate,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    rule = ExtractionRule(
        owner_id=user.feishu_open_id,
        name=body.name,
        sectors=body.sectors,
        fields=[f.model_dump() for f in body.fields],
        prompt_hint=body.prompt_hint,
    )
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return rule


@router.put("/{rule_id}", response_model=ExtractionRuleOut)
async def update_rule(
    rule_id: int,
    body: ExtractionRuleUpdate,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    result = await db.execute(
        select(ExtractionRule).where(
            ExtractionRule.id == rule_id,
            ExtractionRule.owner_id == user.feishu_open_id,
        )
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(404, "规则不存在")
    if body.name is not None:
        rule.name = body.name
    if body.sectors is not None:
        rule.sectors = body.sectors
    if body.fields is not None:
        rule.fields = [f.model_dump() for f in body.fields]
    if body.prompt_hint is not None:
        rule.prompt_hint = body.prompt_hint
    if body.is_active is not None:
        rule.is_active = body.is_active
    await db.commit()
    await db.refresh(rule)
    return rule


@router.delete("/{rule_id}")
async def delete_rule(
    rule_id: int,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    result = await db.execute(
        select(ExtractionRule).where(
            ExtractionRule.id == rule_id,
            ExtractionRule.owner_id == user.feishu_open_id,
        )
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(404, "规则不存在")
    await db.delete(rule)
    await db.commit()
    return {"ok": True}


@router.get("/{rule_id}/data", response_model=ExtractionDataResponse, summary="规则字段汇总数据")
async def get_rule_data(
    rule_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    search: str = Query("", description="在提取字段值中搜索"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> ExtractionDataResponse:
    """获取某个提取规则下所有文档和沟通记录的 key_info 汇总数据。"""
    # 加载规则
    result = await db.execute(
        select(ExtractionRule).where(
            ExtractionRule.id == rule_id,
            ExtractionRule.owner_id == user.feishu_open_id,
        )
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(404, "规则不存在")

    visible_ids = await get_visible_owner_ids(user, db, request)

    # 构建文档查询
    doc_q = select(
        Document.id.label("source_id"),
        literal("document").label("source_type"),
        Document.title.label("source_title"),
        Document.key_info.label("key_info"),
    ).where(
        Document.extraction_rule_id == rule_id,
        Document.key_info.isnot(None),
    )
    if visible_ids is not None:
        doc_q = doc_q.where(Document.owner_id.in_(visible_ids))

    # 构建沟通记录查询
    comm_q = select(
        Communication.id.label("source_id"),
        literal("communication").label("source_type"),
        Communication.title.label("source_title"),
        Communication.key_info.label("key_info"),
    ).where(
        Communication.extraction_rule_id == rule_id,
        Communication.key_info.isnot(None),
    )
    if visible_ids is not None:
        comm_q = comm_q.where(Communication.owner_id.in_(visible_ids))

    # 构建结构化表格查询
    table_q = select(
        StructuredTable.id.label("source_id"),
        literal("structured_table").label("source_type"),
        StructuredTable.name.label("source_title"),
        StructuredTable.key_info.label("key_info"),
    ).where(
        StructuredTable.extraction_rule_id == rule_id,
        StructuredTable.key_info.isnot(None),
    )
    if visible_ids is not None:
        table_q = table_q.where(StructuredTable.owner_id.in_(visible_ids))

    # 搜索过滤
    if search:
        like = f"%{search}%"
        doc_q = doc_q.where(
            Document.title.ilike(like) | cast(Document.key_info, String).ilike(like)
        )
        comm_q = comm_q.where(
            Communication.title.ilike(like) | cast(Communication.key_info, String).ilike(like)
        )
        table_q = table_q.where(
            StructuredTable.name.ilike(like) | cast(StructuredTable.key_info, String).ilike(like)
        )

    # 合并查询
    combined = union_all(doc_q, comm_q, table_q).subquery()

    # 总数
    total = (await db.execute(select(func.count()).select_from(combined))).scalar() or 0

    # 分页数据
    items_stmt = select(combined).offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(items_stmt)).all()

    items = [
        ExtractionDataItem(
            source_type=row.source_type,
            source_id=row.source_id,
            source_title=row.source_title or "",
            key_info=row.key_info or {},
        )
        for row in rows
    ]

    return ExtractionDataResponse(
        rule=ExtractionRuleOut.model_validate(rule),
        items=items,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/{rule_id}/export", summary="导出规则字段汇总为 XLSX")
async def export_rule_data(
    rule_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    search: str = Query("", description="在提取字段值中搜索"),
):
    """导出某个提取规则下所有提取数据为 Excel。"""
    from io import BytesIO
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # 加载规则
    result = await db.execute(
        select(ExtractionRule).where(
            ExtractionRule.id == rule_id,
            ExtractionRule.owner_id == user.feishu_open_id,
        )
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(404, "规则不存在")

    visible_ids = await get_visible_owner_ids(user, db, request)
    fields = rule.fields or []

    # 构建查询（复用 get_rule_data 的逻辑，但不分页）
    doc_q = select(
        Document.id.label("source_id"),
        literal("document").label("source_type"),
        Document.title.label("source_title"),
        Document.key_info.label("key_info"),
    ).where(
        Document.extraction_rule_id == rule_id,
        Document.key_info.isnot(None),
    )
    if visible_ids is not None:
        doc_q = doc_q.where(Document.owner_id.in_(visible_ids))

    comm_q = select(
        Communication.id.label("source_id"),
        literal("communication").label("source_type"),
        Communication.title.label("source_title"),
        Communication.key_info.label("key_info"),
    ).where(
        Communication.extraction_rule_id == rule_id,
        Communication.key_info.isnot(None),
    )
    if visible_ids is not None:
        comm_q = comm_q.where(Communication.owner_id.in_(visible_ids))

    table_q = select(
        StructuredTable.id.label("source_id"),
        literal("structured_table").label("source_type"),
        StructuredTable.name.label("source_title"),
        StructuredTable.key_info.label("key_info"),
    ).where(
        StructuredTable.extraction_rule_id == rule_id,
        StructuredTable.key_info.isnot(None),
    )
    if visible_ids is not None:
        table_q = table_q.where(StructuredTable.owner_id.in_(visible_ids))

    if search:
        like = f"%{search}%"
        doc_q = doc_q.where(
            Document.title.ilike(like) | cast(Document.key_info, String).ilike(like)
        )
        comm_q = comm_q.where(
            Communication.title.ilike(like) | cast(Communication.key_info, String).ilike(like)
        )
        table_q = table_q.where(
            StructuredTable.name.ilike(like) | cast(StructuredTable.key_info, String).ilike(like)
        )

    combined = union_all(doc_q, comm_q, table_q).subquery()
    rows = (await db.execute(select(combined))).all()

    # 构建列头：来源 + 规则字段
    col_labels = ["来源", "来源类型"] + [f.get("label", f.get("key", "")) for f in fields]

    wb = Workbook()
    ws = wb.active
    ws.title = "提取数据"

    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, label in enumerate(col_labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font

    source_type_labels = {"document": "文档", "communication": "沟通记录", "structured_table": "结构化表格"}
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.source_title or "")
        ws.cell(row=row_idx, column=2, value=source_type_labels.get(row.source_type, row.source_type))
        key_info = row.key_info or {}
        for col_offset, f in enumerate(fields):
            val = key_info.get(f.get("label", "")) or key_info.get(f.get("key", ""), "")
            if isinstance(val, (list, dict)):
                val = str(val)
            ws.cell(row=row_idx, column=3 + col_offset, value=val)

    # 自动列宽
    for col_idx, label in enumerate(col_labels, 1):
        max_len = len(str(label)) * 2
        for r in range(2, len(rows) + 2):
            cell_val = ws.cell(row=r, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{rule.name}_提取数据.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/templates")
async def get_templates():
    return {
        "sectors": SECTOR_LABELS,
        "templates": {k: v for k, v in SECTOR_TEMPLATES.items() if k != "common"},
        "common_fields": SECTOR_TEMPLATES["common"],
        "get_fields_example": get_template_fields(["energy"]),
    }
